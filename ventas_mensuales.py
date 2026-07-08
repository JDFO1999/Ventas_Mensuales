import os
import sys
import time
import calendar
from collections import defaultdict
from datetime import date

import pyodbc
from dbfread import DBF
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

SQL_SERVER = "10.10.70.160"
SQL_DATABASE = "Sistemas"
SQL_USER = "Sa"
SQL_PASSWORD = "Alkosto123"

OUTPUT_DIR = r"C:\Users\Alkosto\Desktop\excel - automatico"
OUTPUT_FILE = "Ventas_Mensuales"

MESES_ES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}

BLUE_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="FFFFFFFF")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
DATA_FONT = Font(name="Calibri", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(horizontal="center")
TITLE_ALIGN = Alignment(horizontal="center")
def thin_blue_top():
    return Border(top=Side(style="thin", color="FF4472C4"))

def thin_black_left():
    return Border(left=Side(style="thin", color="FF000000"))

def thin_black_right():
    return Border(right=Side(style="thin", color="FF000000"))

SECTION_COLORS = {
    "A": "FFD9E1F2",
    "F": "FFFFE699",
    "H": "FFC6E0B4",
    "V": "FFF8CBAD",
}


def conectar_sql():
    conn = pyodbc.connect(
        f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={SQL_SERVER};"
        f"DATABASE={SQL_DATABASE};UID={SQL_USER};PWD={SQL_PASSWORD}",
        timeout=10
    )
    return conn


def crear_tabla_posventas(cursor):
    cursor.execute("IF OBJECT_ID('PosVentas','U') IS NOT NULL DROP TABLE PosVentas")
    cursor.connection.commit()
    cursor.execute("""
        IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'PosVentas')
        CREATE TABLE PosVentas (
            Fecha      DATE NOT NULL,
            Hora       TINYINT NOT NULL,
            Tienda     VARCHAR(6) NOT NULL,
            Caja       INT NOT NULL,
            Tipo       CHAR(2) NOT NULL,
            STipo      CHAR(2) NULL,
            Numero     VARCHAR(15) NOT NULL,
            Anulada    CHAR(1) NULL,
            Operador   VARCHAR(20) NULL,
            SubTotal   DECIMAL(18,2) NULL,
            MontoUSD   DECIMAL(18,2) NOT NULL,
            MontoBS    DECIMAL(18,2) NULL,
            Impuesto   DECIMAL(18,2) NULL,
            Descuento  DECIMAL(18,2) NULL,
            IGTF       DECIMAL(18,2) NULL,
            Redondeo   DECIMAL(18,2) NULL,
            TasaDOL    DECIMAL(20,6) NULL,
            CodCli     VARCHAR(10) NULL,
            Nombres    VARCHAR(100) NULL,
            Apellidos  VARCHAR(100) NULL,
            NIT        VARCHAR(14) NULL,
            RIF        VARCHAR(14) NULL,
            CodVen     VARCHAR(2) NULL,
            NroZ       VARCHAR(10) NULL,
            Credito    BIT NULL,
            Vuelto     DECIMAL(16,2) NULL,
            LIMPRIMIO  BIT NULL,
            NroCie     VARCHAR(8) NULL,
            FechaCie   DATE NULL,
            FechaCarga DATETIME DEFAULT GETDATE(),
            CONSTRAINT PK_PosVentas PRIMARY KEY (Tienda, Caja, Numero, Tipo)
        )
    """)
    cursor.connection.commit()

    cursor.execute("""
        IF NOT EXISTS (SELECT * FROM sys.indexes WHERE name = 'IX_PosVentas_FechaTienda')
        CREATE INDEX IX_PosVentas_FechaTienda ON PosVentas(Fecha, Tienda)
    """)
    cursor.connection.commit()


def leer_desde_sql(cursor, year, month):
    cursor.execute("""
        SELECT 
            Tienda,
            Hora,
            SUM(CASE WHEN Tipo = 'FA' THEN MontoUSD ELSE -MontoUSD END) as TotalUSD,
            COUNT(CASE WHEN Tipo = 'FA' THEN 1 END) as Facturas
        FROM PosVentas
        WHERE YEAR(Fecha) = ? AND MONTH(Fecha) = ?
        GROUP BY Tienda, Hora
        ORDER BY Tienda, Hora
    """, (year, month))

    data = {}
    for row in cursor.fetchall():
        tienda = row.Tienda.strip()
        hora = str(row.Hora).zfill(2)
        monto = float(row.TotalUSD)
        facturas = int(row.Facturas)

        if tienda not in data:
            data[tienda] = {"ventas": defaultdict(float), "conteo": defaultdict(int), "total": 0, "facturas": 0}
        
        data[tienda]["ventas"][hora] += monto
        data[tienda]["conteo"][hora] += facturas
        data[tienda]["total"] += monto
        data[tienda]["facturas"] += facturas

    return data


def obtener_sucursales(cursor):
    cursor.execute("""
        SELECT cCodigo, cNombre, cRutaIP, cRutaDBF, cCia
        FROM Sucursal
        WHERE lInactiva = 0 OR lInactiva IS NULL
        ORDER BY cCodigo
    """)
    sucursales = []
    for row in cursor.fetchall():
        sucursales.append({
            "codigo": row.cCodigo.strip() if row.cCodigo else "",
            "nombre": row.cNombre.strip() if row.cNombre else "",
            "ruta_ip": row.cRutaIP.split()[0] if row.cRutaIP and row.cRutaIP.strip() else "",
            "ruta_dbf": row.cRutaDBF.strip() if row.cRutaDBF else "",
            "cia": row.cCia.strip() if row.cCia else "",
        })
    return sucursales


def listar_pos(codigo, ruta_ip, ruta_dbf, modo):
    if modo == "S":
        if ruta_dbf:
            path = f"{ruta_dbf.rstrip('\\')}\\CIERRE_POS"
        else:
            path = f"S:\\aBC-Soft\\Data\\{codigo.rstrip()}\\CIERRE_POS"
    else:
        path = f"\\\\{ruta_ip.rstrip()}\\Sistema\\aBC-Soft\\Cierre_POS"
    path = path.rstrip()

    if not os.path.exists(path):
        if modo == "S":
            return [], f"Drive S: no accesible. Verifique que S: este mapeado. Path: {path}"
        else:
            return [], f"Ruta no existe: {path}"
    pos_dirs = []
    error_msg = None
    try:
        for entry in os.listdir(path):
            full = os.path.join(path, entry)
            if os.path.isdir(full) and entry.upper().startswith("POS"):
                try:
                    pos_num = int(entry[3:])
                    pos_dirs.append((pos_num, full))
                except ValueError:
                    pass
    except PermissionError:
        error_msg = "Permiso denegado"
    except FileNotFoundError:
        error_msg = "Ruta desaparecio durante lectura"
    except TimeoutError:
        error_msg = "Timeout de red"
    except OSError as e:
        error_msg = f"Error de red ({e})"
    if error_msg and not pos_dirs:
        return [], error_msg
    if not pos_dirs:
        return [], f"No se encontraron carpetas POS en {path}"
    return sorted(pos_dirs), None


def leer_fc_archivo(filepath, year_objetivo, month_objetivo, codigo_tienda=None, pos_num=None):
    ventas_por_hora = defaultdict(float)
    conteo_por_hora = defaultdict(int)
    registros_sql = []

    try:
        table = DBF(filepath, encoding="latin-1", char_decode_errors="replace")
    except Exception as e:
        print(f"      ERROR DBF: {e}")
        return {}, 0, defaultdict(float), defaultdict(int), []

    for rec in table:
        try:
            if rec.get("ANULADA", "F") == "T":
                continue

            tipo = rec.get("TIPO", "")
            if tipo not in ("FA", "DV"):
                continue

            fecha = rec.get("FECHA")
            if fecha is None or not isinstance(fecha, date):
                continue

            if fecha.year != year_objetivo or fecha.month != month_objetivo:
                continue

            hora24 = rec.get("HORA24", "")
            if not hora24 or len(hora24) < 2:
                continue

            hora = hora24[:2]

            monto_usd = rec.get("MONTOFACDL", 0) or 0

            if tipo == "FA":
                ventas_por_hora[hora] += monto_usd
                conteo_por_hora[hora] += 1
            elif tipo == "DV":
                ventas_por_hora[hora] -= monto_usd

            if codigo_tienda is not None and pos_num is not None:
                numero = rec.get("NUMERO", "") or ""
                registros_sql.append((
                    fecha, int(hora), codigo_tienda, pos_num,
                    tipo,
                    rec.get("STIPO", "") or "",
                    numero.strip(),
                    rec.get("ANULADA", "") or "",
                    rec.get("OPERADOR", "") or "",
                    rec.get("SUBTOTAL", 0) or 0,
                    monto_usd,
                    rec.get("MONTOFAC", 0) or 0,
                    rec.get("IMPUESTO", 0) or 0,
                    rec.get("DESCUENTO", 0) or 0,
                    rec.get("IGTF", 0) or 0,
                    rec.get("REDONDEO", 0) or 0,
                    rec.get("TASADOL", 0) or 0,
                    rec.get("CODCLI", "") or "",
                    rec.get("NOMBRES", "") or "",
                    rec.get("APELLIDOS", "") or "",
                    rec.get("NIT", "") or "",
                    rec.get("RIF", "") or "",
                    rec.get("CODVEN", "") or "",
                    rec.get("NRO_Z", "") or "",
                    1 if rec.get("CREDITO") else 0,
                    rec.get("VUELTO", 0) or 0,
                    1 if rec.get("LIMPRIMIO") else 0,
                    rec.get("NRO_CIE", "") or "",
                    rec.get("DATE_CIE"),
                ))
        except Exception:
            pass

    total_facturas = sum(conteo_por_hora.values())
    total_ventas = sum(ventas_por_hora.values())

    return ventas_por_hora, total_ventas, conteo_por_hora, total_facturas, registros_sql


def formato_hora(hora_str):
    hora = int(hora_str)
    if hora == 0:
        return "12:00 AM"
    elif hora < 12:
        return f"{hora:02d}:00 AM"
    elif hora == 12:
        return "12:00 PM"
    else:
        return f"{hora - 12:02d}:00 PM"


def generar_excel(resultados, year, month, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    nombre_mes = MESES_ES.get(month, f"MES-{month:02d}")
    ws.title = nombre_mes

    # Title row - merged B2:K3
    ws.merge_cells("B2:K3")
    title_cell = ws["B2"]
    title_cell.value = f"{nombre_mes} ({year})"
    title_cell.font = TITLE_FONT
    title_cell.fill = BLUE_FILL
    title_cell.alignment = TITLE_ALIGN
    ws.row_dimensions[2].height = 12.75
    ws.row_dimensions[3].height = 12.75

    for col_letter in ["C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws[f"{col_letter}2"].fill = BLUE_FILL
        ws[f"{col_letter}3"].fill = BLUE_FILL

    # Headers - row 4
    headers = [
        "TIENDA",
        "Promedio por factura",
        "Clientes atendidos en el mes",
        "Cliente promedio por d\u00eda",
        "Hora con Mayor movimiento",
        "Clientes atendidos con mayor movimiento",
        "Hora con menor moviento",
        "Clientes atendidos con menor movimiento",
        "Total",
        "Proemdio venta por d\u00eda (30)",
    ]
    ws.row_dimensions[4].height = 39.75

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx + 1, value=header)
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.alignment = HEADER_ALIGN

    column_widths = {
        "B": 9.43, "C": 10.57, "D": 17.71, "E": 16.43,
        "F": 38.14, "G": 20.38, "H": 15.0, "I": 21.43,
        "J": 13.29, "K": 15.14,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    row = 5
    dias_mes = calendar.monthrange(year, month)[1]

    for res in resultados:
        tienda = res["tienda"]
        promedio_factura = res["promedio_factura"]
        clientes_mes = res["clientes"]
        total = res["total"]

        formula_promedio_cliente = f"=D{row}/{dias_mes}"
        formula_promedio_venta = f"=J{row}/{dias_mes}"

        hora_mayor = res["hora_mayor"]
        clientes_mayor = res["clientes_mayor"]
        hora_menor = res["hora_menor"]
        clientes_menor = res["clientes_menor"]

        data = [
            tienda,
            promedio_factura,
            clientes_mes,
            formula_promedio_cliente,
            hora_mayor,
            clientes_mayor,
            hora_menor,
            clientes_menor,
            total,
            formula_promedio_venta,
        ]

        section_fill_color = SECTION_COLORS.get(tienda, "00000000")
        section_fill = PatternFill(start_color=section_fill_color, end_color=section_fill_color, fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            top=Side(style="thin"),
            right=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col_idx + 1, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.fill = section_fill
            cell.border = thin_border

            col_letter = chr(ord("A") + col_idx)
            if col_letter == "B":
                cell.number_format = "General"
            elif col_letter == "C":
                cell.number_format = "0.00"
            elif col_letter in ("D", "E", "G", "I"):
                cell.number_format = "#,##0"
            elif col_letter in ("F", "H"):
                cell.number_format = "General"
            elif col_letter in ("J", "K"):
                cell.number_format = "#,##0.00"

        ws.row_dimensions[row].height = 15.0

        row += 1

    try:
        if os.path.exists(output_path):
            os.remove(output_path)
        wb.save(output_path)
        print(f"\nArchivo guardado: {output_path}")
    except PermissionError:
        try:
            alt_path = os.path.join(OUTPUT_DIR, f"Ventas_Mensuales_{nombre_mes}_{year}.xlsx")
            if os.path.exists(alt_path):
                os.remove(alt_path)
            wb.save(alt_path)
            print(f"\nArchivo guardado: {alt_path}")
        except PermissionError:
            print(f"\nERROR: No se pudo guardar el archivo. Cierre Excel e intente de nuevo.")


def main():
    t_total_start = time.time()

    print("=" * 60)
    print("  GENERADOR DE VENTAS MENSUALES DESDE POS")
    print("=" * 60)

    print("\nConectando a SQL Server...")
    try:
        conn = conectar_sql()
        cursor = conn.cursor()
        print("  Conexion exitosa.")
    except Exception as e:
        print(f"  ERROR de conexion: {e}")
        sys.exit(1)

    print("\nObteniendo lista de sucursales...")
    sucursales = obtener_sucursales(cursor)
    conn.close()
    print(f"  {len(sucursales)} sucursales activas encontradas.")

    print("\n" + "-" * 40)
    try:
        mes_input = input("Ingrese el numero de MES (1-12): ").strip()
        mes = int(mes_input)
        if mes < 1 or mes > 12:
            print("  ERROR: Mes fuera de rango. Debe ser 1-12.")
            sys.exit(1)

        anio_input = input("Ingrese el A#o (ej. 2026): ").strip()
        anio = int(anio_input)
        if anio < 2000 or anio > 2100:
            print("  ERROR: A#o fuera de rango.")
            sys.exit(1)
    except ValueError:
        print("  ERROR: Debe ingresar numeros validos.")
        sys.exit(1)
    except KeyboardInterrupt:
        print("\n  Cancelado por el usuario.")
        sys.exit(0)

    nombre_mes = MESES_ES.get(mes, f"MES-{mes:02d}")
    print(f"\n  Procesando: {nombre_mes} {anio}")
    print("-" * 40)

    # Modo de lectura
    print("\nModo de lectura:")
    print("  [1] Servidor S: (S:\\aBC-Soft\\Data\\{codigo}\\CIERRE_POS)")
    print("  [2] Tiendas directo (\\\\IP\\Sistema\\aBC-Soft\\Cierre_POS)")
    try:
        modo_input = input("Seleccione (1/2): ").strip()
        if modo_input == "1":
            modo = "S"
            print(f"  Modo: Servidor S:")
        elif modo_input == "2":
            modo = "IP"
            print(f"  Modo: Tiendas IP")
        else:
            print("  Opcion invalida. Usando modo Tiendas IP por defecto.")
            modo = "IP"
    except (KeyboardInterrupt, EOFError):
        print("\n  Cancelado.")
        sys.exit(0)

    print("-" * 40)

    total_sucursales = len(sucursales)

    print("\nSeleccion de tiendas:")
    print("  [1] Todas las tiendas")
    print("  [2] Elegir manualmente")
    try:
        sel_input = input("Seleccione (1/2): ").strip()
    except (KeyboardInterrupt, EOFError):
        print("\n  Cancelado.")
        sys.exit(0)

    if sel_input == "2":
        print(f"\n  Tiendas disponibles ({total_sucursales}):")
        for i, s in enumerate(sucursales):
            print(f"    [{i+1:2d}] {s['codigo']:6s} - {s['nombre']}")
        try:
            seleccion = input("\n  Ingrese numeros separados por comas (ej: 1,3,5): ").strip()
        except (KeyboardInterrupt, EOFError):
            print("\n  Cancelado.")
            sys.exit(0)

        indices = []
        for parte in seleccion.split(","):
            parte = parte.strip()
            if "-" in parte:
                try:
                    ini, fin = parte.split("-", 1)
                    indices.extend(range(int(ini.strip()), int(fin.strip()) + 1))
                except ValueError:
                    pass
            else:
                try:
                    indices.append(int(parte))
                except ValueError:
                    pass

        sucursales = [sucursales[i - 1] for i in indices if 1 <= i <= len(sucursales)]
        print(f"\n  Seleccionadas: {len(sucursales)} tiendas.")
    else:
        print(f"\n  Procesando todas: {total_sucursales} tiendas.")

    print("-" * 40)

    save_sucursales = sucursales
    resultados = [None] * len(save_sucursales)
    faltantes = []
    total_sucursales = len(sucursales)

    # Conectar SQL, crear tabla
    conn = conectar_sql()
    cursor = conn.cursor()
    crear_tabla_posventas(cursor)

    # Verificar tienda por tienda cuales ya tienen datos en SQL
    tiendas_con_sql = set()
    for sucursal in sucursales:
        codigo = sucursal["codigo"]
        cursor.execute(
            "SELECT COUNT(*) FROM PosVentas WHERE Tienda=? AND YEAR(Fecha)=? AND MONTH(Fecha)=?",
            (codigo, anio, mes)
        )
        if cursor.fetchone()[0] > 0:
            tiendas_con_sql.add(codigo)

    if tiendas_con_sql:
        sql_nombres = [c for c in tiendas_con_sql]
        print(f"\n  {len(tiendas_con_sql)} tiendas con datos en SQL: {', '.join(sql_nombres)}")
        print("  Estas se leeran desde la base de datos.")

    tiendas_sin_sql = [s for s in sucursales if s["codigo"] not in tiendas_con_sql]
    if tiendas_sin_sql:
        sin_nombres = [s["codigo"] for s in tiendas_sin_sql]
        print(f"  {len(tiendas_sin_sql)} tiendas sin datos en SQL: {', '.join(sin_nombres)}")
        print("  Estas se leeran desde archivos DBF.")

    conn.close()

    dias_mes = calendar.monthrange(anio, mes)[1]

    # --- Procesar tiendas DESDE SQL ---
    if tiendas_con_sql:
        print(f"\n--- Leyendo {len(tiendas_con_sql)} tiendas desde SQL ---")
        conn = conectar_sql()
        cursor = conn.cursor()
        data_sql = leer_desde_sql(cursor, anio, mes)
        conn.close()

        for sucursal in sucursales:
            if sucursal["codigo"] not in tiendas_con_sql:
                continue
            codigo = sucursal["codigo"]
            tienda_letra = codigo[0].upper() if codigo else "?"
            sd = data_sql.get(codigo, {})

            clientes = sd.get("facturas", 0)
            total = sd.get("total", 0)
            conteo = sd.get("conteo", defaultdict(int))
            promedio = total / clientes if clientes > 0 else 0

            if conteo:
                hmk = max(conteo, key=conteo.get)
                cm = conteo[hmk]; hm = formato_hora(hmk)
                hmink = min(conteo, key=conteo.get)
                cmi = conteo[hmink]; hmi = formato_hora(hmink)
            else:
                hm = ""; cm = 0; hmi = ""; cmi = 0

            idx_res = save_sucursales.index(sucursal)
            resultados[idx_res] = {
                "tienda": tienda_letra, "promedio_factura": promedio,
                "clientes": clientes, "total": total,
                "hora_mayor": hm, "clientes_mayor": cm,
                "hora_menor": hmi, "clientes_menor": cmi,
            }

    # --- Procesar tiendas DESDE DBF ---
    if tiendas_sin_sql:
        print(f"\n--- Leyendo {len(tiendas_sin_sql)} tiendas desde DBF ---")
        print("-" * 40)

        conn_sql = conectar_sql()
        cursor_sql = conn_sql.cursor()
        cursor_sql.fast_executemany = True
        crear_tabla_posventas(cursor_sql)

        sucursales = tiendas_sin_sql

    for idx, sucursal in enumerate(sucursales):
        t_store_start = time.time()
        codigo = sucursal["codigo"]
        nombre_tienda = sucursal["nombre"]
        ruta_ip = sucursal["ruta_ip"]
        ruta_dbf = sucursal["ruta_dbf"]
        tienda_letra = codigo[0].upper() if codigo else "?"

        print(f"\n[{idx+1}/{total_sucursales}] {codigo} - {nombre_tienda}")
        sys.stdout.flush()

        if modo == "IP" and not ruta_ip:
            print(f"  FALLO: Sin IP/ruta configurada en la base de datos.")
            faltantes.append(sucursal)
            idx_res = save_sucursales.index(sucursal)
            resultados[idx_res] = {
                "tienda": tienda_letra,
                "promedio_factura": 0,
                "clientes": 0,
                "total": 0,
                "hora_mayor": "",
                "clientes_mayor": 0,
                "hora_menor": "",
                "clientes_menor": 0,
            }
            continue

        pos_dirs, error_listado = listar_pos(codigo, ruta_ip, ruta_dbf, modo)
        if not pos_dirs and modo == "S" and ruta_ip:
            print(f"  S: fallo, intentando por IP...")
            pos_dirs, error_listado = listar_pos(codigo, ruta_ip, ruta_dbf, "IP")

        if not pos_dirs:
            razon = error_listado or "Motivo desconocido"
            print(f"  FALLO: {razon}")
            faltantes.append(sucursal)
            idx_res = save_sucursales.index(sucursal)
            resultados[idx_res] = {
                "tienda": tienda_letra,
                "promedio_factura": 0,
                "clientes": 0,
                "total": 0,
                "hora_mayor": "",
                "clientes_mayor": 0,
                "hora_menor": "",
                "clientes_menor": 0,
            }
            continue

        ventas_combinadas = defaultdict(float)
        conteo_combinado = defaultdict(int)
        cajas_procesadas = 0
        store_regs = []
        total_pos = len(pos_dirs)

        for i, (pos_num, pos_path) in enumerate(pos_dirs):
            dbf_filename = f"FC{pos_num:02d}{anio}.DBF"
            dbf_path = os.path.join(pos_path, dbf_filename)

            if not os.path.exists(dbf_path):
                continue

            try:
                size_mb = os.path.getsize(dbf_path) / (1024 * 1024)
            except OSError:
                size_mb = 0

            print(f"  [{i+1}/{total_pos}] {dbf_filename} ({size_mb:.1f} MB)", end="")
            sys.stdout.flush()

            t_file_start = time.time()
            vph, total_ventas, cph, total_facturas, regs_sql = leer_fc_archivo(
                dbf_path, anio, mes, codigo, pos_num
            )
            if regs_sql:
                store_regs.extend(regs_sql)
            t_file_elapsed = time.time() - t_file_start

            if total_facturas > 0:
                for h, monto in vph.items():
                    ventas_combinadas[h] += monto
                for h, count in cph.items():
                    conteo_combinado[h] += count
                cajas_procesadas += 1
                print(f" -> {total_facturas} facts, ${total_ventas:,.0f} ({t_file_elapsed:.1f}s)")
            else:
                print(f" -> sin datos ({t_file_elapsed:.1f}s)")

            sys.stdout.flush()

        total_ventas = sum(ventas_combinadas.values())
        total_facturas = sum(conteo_combinado.values())
        t_store_elapsed = time.time() - t_store_start

        if cajas_procesadas == 0:
            print(f"  Sin datos de ventas para {nombre_mes} {anio}. ({t_store_elapsed:.0f}s)")
            idx_res = save_sucursales.index(sucursal)
            resultados[idx_res] = {
                "tienda": tienda_letra,
                "promedio_factura": 0,
                "clientes": 0,
                "total": 0,
                "hora_mayor": "",
                "clientes_mayor": 0,
                "hora_menor": "",
                "clientes_menor": 0,
            }
            continue

        promedio_factura = total_ventas / total_facturas if total_facturas > 0 else 0

        if conteo_combinado:
            hora_mayor_key = max(conteo_combinado, key=conteo_combinado.get)
            clientes_mayor = conteo_combinado[hora_mayor_key]
            hora_mayor = formato_hora(hora_mayor_key)
            hora_menor_key = min(conteo_combinado, key=conteo_combinado.get)
            clientes_menor = conteo_combinado[hora_menor_key]
            hora_menor = formato_hora(hora_menor_key)
        else:
            hora_mayor = ""
            clientes_mayor = 0
            hora_menor = ""
            clientes_menor = 0

        print(f"  => {cajas_procesadas} cajas | {total_facturas:,.0f} facts | Total: ${total_ventas:,.2f} | Prom: ${promedio_factura:,.2f} | Pico: {hora_mayor} ({clientes_mayor}) | ({t_store_elapsed:.0f}s)")
        sys.stdout.flush()

        # Guardar en SQL Server para esta tienda (resume: si ya tiene datos, saltar)
        cursor_sql.execute("SELECT COUNT(*) FROM PosVentas WHERE Tienda=? AND YEAR(Fecha)=? AND MONTH(Fecha)=?",
            (codigo, anio, mes))
        if cursor_sql.fetchone()[0] == 0 and store_regs:
            # Borrar datos viejos de esta tienda
            cursor_sql.execute("DELETE FROM PosVentas WHERE Tienda=? AND YEAR(Fecha)=? AND MONTH(Fecha)=?",
                (codigo, anio, mes))
            conn_sql.commit()
            # Insertar en chunks
            CHUNK = 5000
            for i in range(0, len(store_regs), CHUNK):
                chunk = store_regs[i:i+CHUNK]
                cursor_sql.executemany(
                    "INSERT INTO PosVentas (Fecha,Hora,Tienda,Caja,Tipo,STipo,Numero,Anulada,Operador,"
                    "SubTotal,MontoUSD,MontoBS,Impuesto,Descuento,IGTF,Redondeo,TasaDOL,"
                    "CodCli,Nombres,Apellidos,NIT,RIF,CodVen,NroZ,Credito,Vuelto,LIMPRIMIO,NroCie,FechaCie) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", chunk)
                conn_sql.commit()
            print(f"  SQL: {len(store_regs)} registros guardados.")

        idx_res = save_sucursales.index(sucursal)
        resultados[idx_res] = {
            "tienda": tienda_letra,
            "promedio_factura": promedio_factura,
            "clientes": total_facturas,
            "total": total_ventas,
            "hora_mayor": hora_mayor,
            "clientes_mayor": clientes_mayor,
            "hora_menor": hora_menor,
            "clientes_menor": clientes_menor,
        }

    t_total_elapsed = time.time() - t_total_start
    print(f"\n{'=' * 60}")
    print(f"Procesamiento completado en {t_total_elapsed/60:.1f} minutos.")

    # Preguntar por tiendas que fallaron
    if faltantes:
        print(f"\n{len(faltantes)} tiendas NO pudieron ser procesadas:")
        for s in faltantes:
            print(f"  - {s['codigo']} ({s['nombre']})")
        try:
            resp = input("\nDesea reintentar estas tiendas? (s/n): ").strip().lower()
        except (KeyboardInterrupt, EOFError):
            resp = "n"

        if resp == "s":
            print(f"\nReintentando {len(faltantes)} tiendas...")
            for sucursal in faltantes:
                t_store_start = time.time()
                codigo = sucursal["codigo"]
                nombre_tienda = sucursal["nombre"]
                ruta_ip = sucursal["ruta_ip"]
                ruta_dbf = sucursal["ruta_dbf"]
                tienda_letra = codigo[0].upper() if codigo else "?"

                print(f"\n  {codigo} - {nombre_tienda}")
                sys.stdout.flush()

                pos_dirs, error_listado = listar_pos(codigo, ruta_ip, ruta_dbf, modo)
                if not pos_dirs and modo == "S" and ruta_ip:
                    print(f"    S: fallo, intentando por IP...")
                    pos_dirs, error_listado = listar_pos(codigo, ruta_ip, ruta_dbf, "IP")

                if not pos_dirs:
                    print(f"    SIGUE FALLANDO: {error_listado or 'Motivo desconocido'}")
                    continue

                ventas_combinadas = defaultdict(float)
                conteo_combinado = defaultdict(int)
                cajas_procesadas = 0
                store_regs = []
                total_pos = len(pos_dirs)

                for i, (pos_num, pos_path) in enumerate(pos_dirs):
                    dbf_filename = f"FC{pos_num:02d}{anio}.DBF"
                    dbf_path = os.path.join(pos_path, dbf_filename)

                    if not os.path.exists(dbf_path):
                        continue

                    try:
                        size_mb = os.path.getsize(dbf_path) / (1024 * 1024)
                    except OSError:
                        size_mb = 0

                    print(f"    [{i+1}/{total_pos}] {dbf_filename} ({size_mb:.1f} MB)", end="")
                    sys.stdout.flush()

                    t_file_start = time.time()
                    vph, total_ventas, cph, total_facturas, regs_sql = leer_fc_archivo(
                        dbf_path, anio, mes, codigo, pos_num
                    )
                    if regs_sql:
                        store_regs.extend(regs_sql)
                    t_file_elapsed = time.time() - t_file_start

                    if total_facturas > 0:
                        for h, monto in vph.items():
                            ventas_combinadas[h] += monto
                        for h, count in cph.items():
                            conteo_combinado[h] += count
                        cajas_procesadas += 1
                        print(f" -> {total_facturas} facts, ${total_ventas:,.0f} ({t_file_elapsed:.1f}s)")
                    else:
                        print(f" -> sin datos ({t_file_elapsed:.1f}s)")

                    sys.stdout.flush()

                if cajas_procesadas == 0:
                    print(f"    Sin datos de ventas. ({time.time() - t_store_start:.0f}s)")
                    continue

                total_ventas = sum(ventas_combinadas.values())
                total_facturas = sum(conteo_combinado.values())
                t_store_elapsed = time.time() - t_store_start

                promedio_factura = total_ventas / total_facturas if total_facturas > 0 else 0

                if conteo_combinado:
                    hora_mayor_key = max(conteo_combinado, key=conteo_combinado.get)
                    clientes_mayor = conteo_combinado[hora_mayor_key]
                    hora_mayor = formato_hora(hora_mayor_key)
                    hora_menor_key = min(conteo_combinado, key=conteo_combinado.get)
                    clientes_menor = conteo_combinado[hora_menor_key]
                    hora_menor = formato_hora(hora_menor_key)
                else:
                    hora_mayor = ""
                    clientes_mayor = 0
                    hora_menor = ""
                    clientes_menor = 0

                print(f"    => {cajas_procesadas} cajas | {total_facturas:,.0f} facts | Total: ${total_ventas:,.2f} | Prom: ${promedio_factura:,.2f} | Pico: {hora_mayor} ({clientes_mayor}) | ({t_store_elapsed:.0f}s)")
                sys.stdout.flush()

                if store_regs:
                    cursor_sql.execute("DELETE FROM PosVentas WHERE Tienda=? AND YEAR(Fecha)=? AND MONTH(Fecha)=?",
                        (codigo, anio, mes))
                    conn_sql.commit()
                    CHUNK = 5000
                    for i in range(0, len(store_regs), CHUNK):
                        chunk = store_regs[i:i+CHUNK]
                        cursor_sql.executemany(
                            "INSERT INTO PosVentas (Fecha,Hora,Tienda,Caja,Tipo,STipo,Numero,Anulada,Operador,"
                            "SubTotal,MontoUSD,MontoBS,Impuesto,Descuento,IGTF,Redondeo,TasaDOL,"
                            "CodCli,Nombres,Apellidos,NIT,RIF,CodVen,NroZ,Credito,Vuelto,LIMPRIMIO,NroCie,FechaCie) "
                            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", chunk)
                        conn_sql.commit()
                    print(f"    SQL: {len(store_regs)} registros guardados.")

                idx_suc = save_sucursales.index(sucursal)
                resultados[idx_suc] = {
                    "tienda": tienda_letra,
                    "promedio_factura": promedio_factura,
                    "clientes": total_facturas,
                    "total": total_ventas,
                    "hora_mayor": hora_mayor,
                    "clientes_mayor": clientes_mayor,
                    "hora_menor": hora_menor,
                    "clientes_menor": clientes_menor,
                }

    # Restaurar lista original
    sucursales = save_sucursales

    if tiendas_sin_sql:
        conn_sql.close()

    total_facturas_general = sum(r["clientes"] for r in resultados if r is not None)
    if total_facturas_general == 0:
        print("\n" + "=" * 60)
        print("ADVERTENCIA: Ninguna tienda tiene datos para este mes/ano.")
        print("Verifique que las tiendas esten accesibles y el mes sea correcto.")
        print("No se genera archivo Excel vacio.")
        print("=" * 60)
        return

    print("\nGenerando archivo Excel...")

    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    output_path = os.path.join(OUTPUT_DIR, f"{OUTPUT_FILE}_{nombre_mes}_{anio}.xlsx")
    # Eliminar viejo si existe
    if os.path.exists(output_path):
        os.remove(output_path)
    generar_excel(resultados, anio, mes, output_path)

    print(f"Tiempo total: {t_total_elapsed/60:.1f} minutos.")
    print("=" * 60)


if __name__ == "__main__":
    main()
